import io
from datetime import datetime
from typing import List, Annotated

from fastapi import APIRouter, Depends, status, File, UploadFile, HTTPException, Query
from fastapi.responses import Response
from openpyxl.reader.excel import load_workbook
from sqlalchemy import delete
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from sqlmodel import select

from app.auth.dependencies import AccessTokenBearer, RoleChecker, get_current_user
from app.db.main import get_session
from app.db.models import Task, WorkType, Voltage, User
from app.errors import TaskNotFound, InsufficientPermission
from app.tasks.dependencies import get_task_or_404
from app.tasks.schemas import TaskRead, TaskCreate, TaskUpdate
from app.tasks.service import TaskService
from app.tasks.utils import get_file_from_database
from app.utils.photo_metadata import photo_metadata

task_router = APIRouter()
task_service = TaskService()
access_token_bearer = AccessTokenBearer()

admin_checker = Depends(RoleChecker(['admin']))
worker_checker = Depends(RoleChecker(['admin', 'worker']))
user_checker = Depends(RoleChecker(['admin', 'user']))
guest_checker = Depends(RoleChecker(['guest']))
all_roles_checker = Depends(RoleChecker(['admin', 'user', 'worker', 'guest']))

VALID_CODE = '202502'
DOWNLOAD_APK_URL = f"https://firebasestorage.googleapis.com/v0/b/dagenergi-b0086.appspot.com/o/apk%2Fapp-release.apk.zip?alt=media&token=248b1700-a781-45d5-99db-44ffe94d7048"


@task_router.get("/download_apk")
async def download_apk(
		code: str = Query(..., min_length=6, max_length=6)
):
	if code != VALID_CODE:
		raise HTTPException(status_code=403, detail="Invalid Code")
	return  {"download_url": DOWNLOAD_APK_URL}


@task_router.get("/", response_model=List[TaskRead], dependencies=[all_roles_checker])
async def get_all_tasks(
		session: AsyncSession = Depends(get_session),
		_: dict = Depends(access_token_bearer)
):
	stmt = select(Task).options(selectinload(Task.worker)).where(Task.is_completed == False).order_by(Task.created_at)
	result = await session.execute(stmt)
	tasks = result.scalars().all()
	return tasks


@task_router.get("/completed", response_model=List[TaskRead], dependencies=[all_roles_checker])
async def get_completed_task(session: Annotated[AsyncSession, Depends(get_session)]):
	stmt = select(Task).options(selectinload(Task.worker)).where(Task.is_completed == True).order_by(Task.created_at)
	result = await session.execute(stmt)
	tasks = result.scalars().all()
	return tasks


@task_router.get("/{task_id}", response_model=TaskRead, dependencies=[worker_checker])
async def get_task(
		task: Task = Depends(get_task_or_404),
		session: AsyncSession = Depends(get_session),
		_: dict = Depends(access_token_bearer)
):
	return task


@task_router.post(
	"/",
	status_code=status.HTTP_201_CREATED,
	response_model=TaskRead,
	dependencies=[worker_checker]
)
async def add_task(
		task_data: TaskCreate,
		worker: User = Depends(get_current_user),
		session: AsyncSession = Depends(get_session)
):
	coordinates = None
	# Utiliser les deux premières photos pour obtenir les coordonnées
	if task_data.photos and len(task_data.photos) >= 2:
		coordinates = photo_metadata.get_coordinate_from_url(task_data.photos[0])
		if not coordinates:
			coordinates = photo_metadata.get_coordinate_from_url(task_data.photos[1])

	task = Task(**task_data.dict(), worker_id=worker.uid)
	task.is_completed = True
	task.completion_date = datetime.now().strftime("%d-%m-%Y %H:%M")
	if coordinates:
		task.latitude = coordinates.latitude
		task.longitude = coordinates.longitude
	session.add(task)
	await session.commit()
	await session.refresh(task)
	return task


@task_router.post(
	"/upload",
	status_code=status.HTTP_201_CREATED,
	response_model=List[TaskRead]
)
async def upload_file(
		uploadFile: UploadFile = File(...),
		session: AsyncSession = Depends(get_session),
):
	uploadFile.file.seek(0)
	content = uploadFile.file.read()
	workbook = load_workbook(io.BytesIO(content))
	sheet = workbook.active
	tasks = []
	for row in range(3, sheet.max_row + 1):
		try:
			new_task = TaskCreate(
				work_type=str(sheet.cell(row=row, column=2).value) if sheet.cell(row=row, column=2).value else None,
				dispatcher_name=str(sheet.cell(row=row, column=3).value) if sheet.cell(row=row, column=3).value else None,
				address=str(sheet.cell(row=row, column=4).value) if sheet.cell(row=row, column=4).value else None,
				planner_date=str(sheet.cell(row=row, column=5).value) if sheet.cell(row=row, column=5).value else None,
				voltage=sheet.cell(row=row, column=7).value if sheet.cell(row=row, column=7).value else None,
				job=str(sheet.cell(row=row, column=8).value) if sheet.cell(row=row, column=8).value else None,
				latitude=None,
				longitude=None,
				photos=[],
				comments=None
			)
		except KeyError as e:
			raise HTTPException(
				status_code=400, detail=f"Missing column in the Excel file: {e}"
			)
		task = await task_service.create_task_from_file(new_task, session)
		tasks.append(task)
	return tasks


@task_router.patch(
	"/{task_id}",
	response_model=TaskRead,
	dependencies=[worker_checker]
)
async def update_task(
		update_data: TaskUpdate,
		task = Depends(get_task_or_404),
		worker = Depends(get_current_user),
		session: AsyncSession = Depends(get_session),
):
	update_data_dict = update_data.model_dump(exclude_unset=True)
	photos = update_data_dict.get("photos")

	# Si des photos sont fournies, on essaie d'en extraire les coordonnées
	if photos and isinstance(photos, list) and len(photos) > 0:
		first_photo = photos[0]
		coordinates = photo_metadata.get_coordinate_from_url(first_photo)
		if coordinates:
			update_data_dict["latitude"] = coordinates.latitude
			update_data_dict["longitude"] = coordinates.longitude

	for key, value in update_data_dict.items():
		setattr(task, key, value)

	task.worker_id = worker.uid
	task.completion_date =  datetime.now().strftime("%d-%m-%Y %H:%M")
	task.is_completed = True

	await session.commit()
	await session.refresh(task)
	return task

@task_router.delete(
	"/clear", status_code=status.HTTP_204_NO_CONTENT,
	dependencies=[admin_checker]
)
async def delete_all_tasks(
		session: AsyncSession = Depends(get_session)
):
	statement = delete(Task)
	await session.execute(statement)
	await session.commit()


@task_router.delete(
	"/{task_id}", status_code=status.HTTP_204_NO_CONTENT, dependencies=[admin_checker]
)
async def delete_task(
		task = Depends(get_current_user),
		session: AsyncSession = Depends(get_session),

):
	await session.delete(task)
	await session.commit()


@task_router.post("/download", status_code=status.HTTP_201_CREATED, dependencies=[all_roles_checker])
async def download(
	session: AsyncSession = Depends(get_session),
):
	tasks = await task_service.get_tasks_completed(session)
	file = get_file_from_database(tasks)
	file_content = file.getvalue()
	headers = {
		'Content-Disposition': 'attachment; filename="Reports.xlsx"',
		"Access-Control-Allow-Origin": "*",
		"Access-Control-Allow-Headers": "*",
		"Access-Control_Allow-Methods": "POST, GET, OPTIONS",
	}
	return Response(content=file_content,
					media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;charset=utf-8",
					headers=headers)
