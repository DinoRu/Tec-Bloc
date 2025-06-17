from io import BytesIO
from typing import List

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, PatternFill

from app.db.models import Task


def draw_report_header(worksheet: Worksheet) -> None:
	headers = [
		"№", "Тип работ", "Диспетчерское наименование ОЭСХ", "Адрес объекта",
		"Дата работ по плану", "Класс напряжения, кВ", "Работы",
		"Дата выполнения", "Широта", "Долгота",
		"фотофиксация 1", "фотофиксация 2", "фотофиксация 3",
		"фотофиксация 4", "фотофиксация 5",
		"Исполнитель", "Комментарий"
	]
	set_header_row(worksheet, headers)
	set_column_sizes(worksheet)


def set_header_row(worksheet: Worksheet, headers: list[str]) -> None:
	alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
	bold_font = Font(bold=True)

	for col, header in enumerate(headers, start=1):
		cell = worksheet.cell(row=1, column=col, value=header)
		cell.alignment = alignment
		cell.font = bold_font

	worksheet.row_dimensions[1].height = 30


def set_column_sizes(worksheet: Worksheet) -> None:
	column_widths = {
		"A": 10, "B": 30, "C": 30, "D": 30, "E": 30,
		"F": 20, "G": 30, "H": 30, "I": 20, "J": 20,
		"K": 50, "L": 50, "M": 50, "N": 50, "O": 50,
		"P": 40, "Q": 60
	}

	for col, width in column_widths.items():
		worksheet.column_dimensions[col].width = width

	worksheet.row_dimensions[2].height = 30


def get_file_from_database(tasks: List[Task]) -> BytesIO:
	workbook = Workbook()
	worksheet = workbook.active
	draw_report_header(worksheet)

	row_num = 2  # On commence à écrire à partir de la ligne 2 (ligne 1 = en-têtes)

	for task in tasks:
		# Colonnes A à J
		data = [
			task.id,
			task.work_type,
			task.dispatcher_name,
			task.address,
			task.planner_date,
			task.voltage,
			task.job,
			task.completion_date,
			task.latitude,
			task.longitude,
		]

		for col_index, value in enumerate(data, start=1):
			cell = worksheet.cell(row=row_num, column=col_index, value=value)
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

		# Colonnes K à O : Photos
		for i in range(5):
			col_index = 11 + i  # K = 11
			if i < len(task.photos):
				cell = worksheet.cell(row=row_num, column=col_index, value=f"фото {i+1}")
				cell.hyperlink = task.photos[i]
				cell.style = "Hyperlink"
				cell.alignment = Alignment(horizontal="center", vertical="center")
				cell.font = Font(color="0000EE", underline="single")
				# cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
			else:
				worksheet.cell(row=row_num, column=col_index, value="")

		# Colonne P: Worker
		cell = worksheet.cell(row=row_num, column=16, value=task.worker.username if task.worker else "")
		cell.alignment = Alignment(horizontal="center", vertical="center")

		# Colonne Q : Commentaires
		cell = worksheet.cell(row=row_num, column=17, value=task.comments)
		cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

		row_num += 1

	buffer = BytesIO()
	workbook.save(buffer)
	buffer.seek(0)
	return buffer